import os
from openpyxl import load_workbook
import config
import errors
import pyperclip
import math


EMOJI_DICTIONARY = {
    "sun": "/ty",
    "circle": "/zhq",
    "celebrate": "\U0001F389",
    'bell': "\U0001F514"
}


EMOJI_NO_LIST = ["1⃣", "2⃣", "3⃣", "4⃣", "5⃣", "6⃣", "7⃣", "8⃣", "9⃣"]


def convert_num_to_emoji(rank_num):
    rank_num_digits = int(math.log(rank_num, 10) + 1)
    res = ""
    for i in range(rank_num_digits-1, -1, -1):
        digit_value = int(rank_num / (10 ** i))
        res += EMOJI_NO_LIST[digit_value - 1]
        rank_num -= digit_value * (10 ** i)
    return res


def print_divider():
    print('------------------------------')


def print_output_divider():
    print('------------ Preview ------------')


def print_end():
    print_divider()
    print("             END")
    print_divider()


def dict_to_display(dict1):
    res = ""
    for key, value in dict1.items():
        res += f"{key}. {value}\n"

    return res


def get_files_to_dict():
    myfiles = [f for f in os.listdir(config.EXCEL_DIRECTORY)
               if (os.path.isfile(os.path.join(config.EXCEL_DIRECTORY, f)) and (f.endswith(".xlsx")))]

    if not myfiles:
        errors.empty_folder_error(config.EXCEL_DIRECTORY)
    res_dict = {i+1: myfiles[i] for i in range(len(myfiles))}
    return res_dict


def load_ws(file_name):
    # Load ws
    wb = load_workbook(f"{config.EXCEL_DIRECTORY}/{file_name}",data_only=True)
    sheets = wb.sheetnames
    ws = None
    if len(sheets) != 1:
        errors.ws_number_error()
    else:
        ws = wb[sheets[0]]

    return ws


def load_file(file_path):
    f = None
    try:
        f = open(file_path, "r", encoding="UTF-8")
    except FileNotFoundError:
        errors.file_not_found_error(file_path)
    return f


def copy_to_clipboard(sentence):
    pyperclip.copy(sentence)
    print("Result has successfully copied to your clipboard!")


def count_chinese_char(str1):
    res = 0
    for letter in str1:
        if '\u4e00' <= letter <= '\u9fa5':
            res += 1
    return res
